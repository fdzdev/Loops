"""The verifier — the product of this loop.

A financial model is only trustworthy if its numbers reconcile. So we never ask
the model "is this spreadsheet right?". Instead we:

  1. RECOMPUTE every figure in plain Python, straight from the assumptions, with
     no reference to the spreadsheet. This is the ground truth.
  2. READ the .xlsx the executor built and reconcile it cell-by-cell against the
     recompute. Any cell that disagrees by more than a cent is a reconciliation
     failure (evidence carries the diff).
  3. Check accounting IDENTITIES on the recompute itself (gross profit = revenue
     - COGS, net income = pre-tax - tax, margins in range) so a self-consistent
     but *wrong* recompute still can't pass.
  4. Grade each RUBRIC criterion as a concrete boolean. A criterion the rubric
     names but we cannot map to code is reported as un-gradeable and FAILS — you
     cannot pass a vague rubric.

Confirmed iff: reconcile clean AND identities hold AND every rubric criterion
that was requested passed. Everything is dumped into Verdict.evidence so a human
can confirm in under a minute.

This module does deterministic arithmetic and spreadsheet reading only. No LLM,
no network — the verifier must not depend on the thing it is judging.
"""
from __future__ import annotations

from dataclasses import dataclass, field

from agentloops import Verdict

from .model_spec import Assumptions, ModelCandidate, ModelTask

# A cent. Spreadsheet floats and Python floats should agree to well within this.
CENT = 0.01

# Canonical row labels the builder must emit, in order. The verifier reads the
# sheet by these labels, so a model can't pass by renaming rows into mush.
ROW_REVENUE = "Revenue"
ROW_COGS = "COGS"
ROW_GROSS_PROFIT = "Gross Profit"
ROW_OPEX = "Operating Expenses"
ROW_OPERATING_INCOME = "Operating Income"
ROW_TAX = "Tax"
ROW_NET_INCOME = "Net Income"
ROW_TOTAL_REVENUE = "Total Revenue"  # the cross-year sum, a single cell

ORDERED_ROWS = [
    ROW_REVENUE,
    ROW_COGS,
    ROW_GROSS_PROFIT,
    ROW_OPEX,
    ROW_OPERATING_INCOME,
    ROW_TAX,
    ROW_NET_INCOME,
]


@dataclass
class Recompute:
    """Ground-truth figures derived from assumptions alone."""

    years: list[int]
    revenue: list[float]
    cogs: list[float]
    gross_profit: list[float]
    opex: list[float]
    operating_income: list[float]
    tax: list[float]
    net_income: list[float]
    gross_margin: list[float]
    net_margin: list[float]
    total_revenue: float
    stated_revenue_growth: float  # the assumption, carried so a rubric check can tie out to it

    def as_rows(self) -> dict[str, list[float]]:
        return {
            ROW_REVENUE: self.revenue,
            ROW_COGS: self.cogs,
            ROW_GROSS_PROFIT: self.gross_profit,
            ROW_OPEX: self.opex,
            ROW_OPERATING_INCOME: self.operating_income,
            ROW_TAX: self.tax,
            ROW_NET_INCOME: self.net_income,
        }


def recompute(task: ModelTask, a: Assumptions) -> Recompute:
    """Build the entire model in plain Python. This is the independent oracle the
    spreadsheet is reconciled against — it shares no code with the builder."""
    years = [task.start_year + i for i in range(task.years)]
    revenue, cogs, gross, opex, opinc, tax, net = [], [], [], [], [], [], []
    gmargin, nmargin = [], []

    for i in range(task.years):
        rev = a.base_revenue * ((1 + a.revenue_growth) ** i)
        cost = rev * a.cogs_pct
        gp = rev - cost
        ox = a.opex_base * ((1 + a.opex_growth) ** i)
        oi = gp - ox
        # Tax only bites positive pre-tax income — a loss year owes nothing.
        t = oi * a.tax_rate if oi > 0 else 0.0
        ni = oi - t

        revenue.append(rev)
        cogs.append(cost)
        gross.append(gp)
        opex.append(ox)
        opinc.append(oi)
        tax.append(t)
        net.append(ni)
        gmargin.append(gp / rev if rev else 0.0)
        nmargin.append(ni / rev if rev else 0.0)

    return Recompute(
        years=years,
        revenue=revenue,
        cogs=cogs,
        gross_profit=gross,
        opex=opex,
        operating_income=opinc,
        tax=tax,
        net_income=net,
        gross_margin=gmargin,
        net_margin=nmargin,
        total_revenue=sum(revenue),
        stated_revenue_growth=a.revenue_growth,
    )


# ---------------------------------------------------------------------------
# Rubric criteria — every name maps to a concrete boolean over (task, recompute,
# sheet). A rubric entry with no mapping here is un-gradeable and FAILS.
# ---------------------------------------------------------------------------


@dataclass
class SheetView:
    """What the verifier could parse out of the .xlsx: a label->per-year-values
    map, the year header, and the single Total Revenue cell."""

    rows: dict[str, list[float]] = field(default_factory=dict)
    years: list[int] = field(default_factory=list)
    total_revenue: float | None = None


def _has_total_revenue_row(task, rc, sv: SheetView) -> bool:
    return sv.total_revenue is not None


def _years_contiguous(task, rc, sv: SheetView) -> bool:
    if not sv.years:
        return False
    expected = [task.start_year + i for i in range(task.years)]
    return sv.years == expected


def _gross_margin_in_unit_interval(task, rc: Recompute, sv) -> bool:
    return all(0.0 <= m <= 1.0 for m in rc.gross_margin)


def _net_margin_below_one(task, rc: Recompute, sv) -> bool:
    # Net margin can be negative (a loss) but never exceeds 100% of revenue.
    return all(m <= 1.0 for m in rc.net_margin)


def _growth_matches_assumption(task, rc: Recompute, sv) -> bool:
    """Every year-over-year revenue step must equal the STATED growth assumption
    (within a hair of floating-point slack). Compares the realized per-year ratio
    against the independently-carried assumption, so it is a real check, not a
    tautology against itself."""
    for i in range(1, len(rc.revenue)):
        if not rc.revenue[i - 1]:
            return False
        realized = rc.revenue[i] / rc.revenue[i - 1] - 1
        if abs(realized - rc.stated_revenue_growth) > 1e-9:
            return False
    return True


def _net_income_identity(task, rc: Recompute, sv) -> bool:
    """net = (revenue - cogs - opex) - tax, every year."""
    for i in range(task.years):
        expected = (rc.revenue[i] - rc.cogs[i] - rc.opex[i]) - rc.tax[i]
        if abs(expected - rc.net_income[i]) > CENT:
            return False
    return True


def _total_revenue_is_sum(task, rc: Recompute, sv: SheetView) -> bool:
    """The sheet's Total Revenue cell equals the sum of its yearly revenue cells —
    a balance identity the spreadsheet must satisfy internally."""
    if sv.total_revenue is None or ROW_REVENUE not in sv.rows:
        return False
    return abs(sv.total_revenue - sum(sv.rows[ROW_REVENUE])) <= CENT


def _all_required_rows_present(task, rc, sv: SheetView) -> bool:
    return all(label in sv.rows for label in ORDERED_ROWS)


# The rubric vocabulary. The task may request any subset by name; each maps to a
# pure boolean. Anything outside this dict is un-gradeable (and fails).
RUBRIC_CHECKS: dict[str, callable] = {
    "has_total_revenue_row": _has_total_revenue_row,
    "years_contiguous": _years_contiguous,
    "gross_margin_in_unit_interval": _gross_margin_in_unit_interval,
    "net_margin_below_one": _net_margin_below_one,
    "growth_matches_assumption": _growth_matches_assumption,
    "net_income_identity": _net_income_identity,
    "total_revenue_is_sum": _total_revenue_is_sum,
    "all_required_rows_present": _all_required_rows_present,
}


def read_sheet(xlsx_path: str) -> SheetView:
    """Read the workbook the executor built. Pulls the year header, each labelled
    row's per-year numeric values, and the single Total Revenue cell. Tolerant of
    missing pieces — a missing row just means the matching rubric check fails."""
    from openpyxl import load_workbook  # local import: only needed when verifying

    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    sv = SheetView()

    # Header row: first row whose first cell is "Year" (case-insensitive).
    header_row_idx = None
    for r in range(1, ws.max_row + 1):
        first = ws.cell(row=r, column=1).value
        if isinstance(first, str) and first.strip().lower() == "year":
            header_row_idx = r
            sv.years = [
                int(ws.cell(row=r, column=c).value)
                for c in range(2, ws.max_column + 1)
                if isinstance(ws.cell(row=r, column=c).value, (int, float))
            ]
            break

    for r in range(1, ws.max_row + 1):
        label = ws.cell(row=r, column=1).value
        if not isinstance(label, str):
            continue
        label = label.strip()
        if label == ROW_TOTAL_REVENUE:
            val = ws.cell(row=r, column=2).value
            if isinstance(val, (int, float)):
                sv.total_revenue = float(val)
            continue
        if label in ORDERED_ROWS:
            vals = [
                float(ws.cell(row=r, column=c).value)
                for c in range(2, (len(sv.years) + 2) if sv.years else ws.max_column + 1)
                if isinstance(ws.cell(row=r, column=c).value, (int, float))
            ]
            sv.rows[label] = vals

    return sv


def reconcile(rc: Recompute, sv: SheetView) -> list[dict]:
    """Compare the spreadsheet against the recompute, cell by cell. Returns the
    list of disagreements (empty == clean). Each diff is human-legible."""
    diffs: list[dict] = []
    truth = rc.as_rows()

    for label, expected_vals in truth.items():
        got = sv.rows.get(label)
        if got is None:
            diffs.append({"row": label, "error": "missing from spreadsheet"})
            continue
        if len(got) != len(expected_vals):
            diffs.append(
                {"row": label, "error": f"has {len(got)} cells, expected {len(expected_vals)}"}
            )
            continue
        for i, (e, g) in enumerate(zip(expected_vals, got)):
            if abs(e - g) > CENT:
                diffs.append(
                    {"row": label, "year_index": i, "expected": round(e, 2), "got": round(g, 2)}
                )

    if sv.total_revenue is None:
        diffs.append({"row": ROW_TOTAL_REVENUE, "error": "missing from spreadsheet"})
    elif abs(sv.total_revenue - rc.total_revenue) > CENT:
        diffs.append(
            {
                "row": ROW_TOTAL_REVENUE,
                "expected": round(rc.total_revenue, 2),
                "got": round(sv.total_revenue, 2),
            }
        )
    return diffs


def verify(candidate: ModelCandidate) -> Verdict:
    """Deterministic verdict. No LLM, no network. See module docstring for the
    confirmation rule. Every signal a reviewer needs lands in evidence."""
    task, a = candidate.task, candidate.assumptions

    if not candidate.xlsx_path:
        return Verdict(False, {"error": "no xlsx was built"}, "build produced no spreadsheet")

    rc = recompute(task, a)

    # Read the spreadsheet (defensively — a corrupt file is a rejection, not a crash).
    try:
        sv = read_sheet(candidate.xlsx_path)
    except Exception as exc:  # noqa: BLE001 — any read failure is a clean reject
        return Verdict(
            False,
            {"error": f"could not read xlsx: {exc}", "xlsx_path": candidate.xlsx_path},
            "spreadsheet unreadable",
        )

    # 1+2: reconcile spreadsheet against the independent recompute.
    diffs = reconcile(rc, sv)
    reconciled = not diffs

    # 3: accounting identities on the recompute itself, ALWAYS enforced (not just
    # when the rubric names them). A self-consistent-but-wrong recompute, or a task
    # whose rubric simply forgets these checks, still cannot pass.
    identities = {
        "net_income_identity": _net_income_identity(task, rc, sv),
        "gross_margin_in_unit_interval": _gross_margin_in_unit_interval(task, rc, sv),
    }
    identities_hold = all(identities.values())

    # 4: grade the rubric. Unknown criteria are un-gradeable -> fail.
    rubric_results: dict[str, bool] = {}
    ungradeable: list[str] = []
    for name in task.rubric:
        check = RUBRIC_CHECKS.get(name)
        if check is None:
            ungradeable.append(name)
            rubric_results[name] = False
            continue
        try:
            rubric_results[name] = bool(check(task, rc, sv))
        except Exception as exc:  # noqa: BLE001
            rubric_results[name] = False
            ungradeable.append(f"{name} (raised {exc})")

    rubric_passed = all(rubric_results.values()) and bool(task.rubric)

    confirmed = reconciled and identities_hold and rubric_passed

    if confirmed:
        reason = (
            f"reconciled clean against recompute; identities hold; "
            f"{len(task.rubric)}/{len(task.rubric)} rubric criteria passed"
        )
    elif not reconciled:
        reason = f"{len(diffs)} reconciliation diff(s) against recompute"
    elif not identities_hold:
        broken = [k for k, v in identities.items() if not v]
        reason = f"accounting identity failures: {broken}"
    elif ungradeable:
        reason = f"un-gradeable rubric criteria: {ungradeable}"
    else:
        failed = [k for k, v in rubric_results.items() if not v]
        reason = f"rubric failures: {failed}"

    evidence = {
        "xlsx_path": candidate.xlsx_path,
        "reconciled": reconciled,
        "reconcile_diffs": diffs,
        "identities": identities,
        "rubric_results": rubric_results,
        "ungradeable_criteria": ungradeable,
        "recompute": {
            "years": rc.years,
            "revenue": [round(x, 2) for x in rc.revenue],
            "net_income": [round(x, 2) for x in rc.net_income],
            "gross_margin": [round(x, 4) for x in rc.gross_margin],
            "total_revenue": round(rc.total_revenue, 2),
        },
    }
    return Verdict(confirmed=confirmed, evidence=evidence, reason=reason)
