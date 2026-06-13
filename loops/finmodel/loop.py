"""finmodel — a rubric-graded financial-model builder loop.

  GENERATOR : a model task spec (title, years, currency, rubric). Each round it
              asks llm.strong to PROPOSE a set of assumptions (base revenue,
              growth, COGS%, opex, tax). One Candidate per proposed build;
              dedupe_key = hash of the assumptions, so the same numbers are never
              re-built or re-verified.
  EXECUTOR  : llm.strong proposes the structure/assumptions as structured data,
              then plain Python builds an .xlsx with openpyxl (the [finmodel]
              extra). The model never writes the figures — code derives them, so
              the spreadsheet and the verifier's recompute share an origin only
              in the assumptions, never in the math.
  VERIFIER  : deterministic. Recompute every figure in plain Python, reconcile the
              spreadsheet against it cell by cell, check accounting identities,
              then grade each rubric criterion as a concrete boolean. See
              verifier.py — it is the product.

Run it:  python -m loops.finmodel          (needs ANTHROPIC_API_KEY + `pip install -e '.[finmodel]'`)
"""
from __future__ import annotations

import os
from typing import Iterable, Optional

from agentloops import LLM, Budget, JsonlState, run_loop

from .model_spec import Assumptions, ModelCandidate, ModelTask
from .verifier import (
    ORDERED_ROWS,
    ROW_COGS,
    ROW_GROSS_PROFIT,
    ROW_NET_INCOME,
    ROW_OPEX,
    ROW_OPERATING_INCOME,
    ROW_REVENUE,
    ROW_TAX,
    ROW_TOTAL_REVENUE,
    verify,
)

# The demo task. Every rubric entry below is a key in verifier.RUBRIC_CHECKS, so
# each one is a concrete boolean — the anti-gaming requirement.
DEMO_TASK = ModelTask(
    title="5-Year Revenue & Net Income Projection",
    currency="USD",
    start_year=2025,
    years=5,
    rubric=[
        "all_required_rows_present",
        "has_total_revenue_row",
        "years_contiguous",
        "gross_margin_in_unit_interval",
        "net_margin_below_one",
        "growth_matches_assumption",
        "net_income_identity",
        "total_revenue_is_sum",
    ],
)


def _derive_rows(task: ModelTask, a: Assumptions):
    """The EXECUTOR's own arithmetic. This deliberately does NOT call the
    verifier's `recompute` — if the builder and the verifier shared the same math,
    reconciliation would be the verifier checking a serialized copy of itself and
    could never catch a builder bug. By deriving figures independently here, a
    mistake in this code surfaces as a real reconciliation diff in `verify`.

    Returns (years, ordered_label->per_year_values, total_revenue)."""
    years = [task.start_year + i for i in range(task.years)]
    revenue, cogs, gross, opex, opinc, tax, net = [], [], [], [], [], [], []

    for i in range(task.years):
        rev = a.base_revenue * ((1 + a.revenue_growth) ** i)
        cost = rev * a.cogs_pct
        gp = rev - cost
        ox = a.opex_base * ((1 + a.opex_growth) ** i)
        oi = gp - ox
        # Tax only bites positive pre-tax income — a loss year owes nothing.
        t = oi * a.tax_rate if oi > 0 else 0.0
        ni = oi - t

        revenue.append(rev)
        cogs.append(cost)
        gross.append(gp)
        opex.append(ox)
        opinc.append(oi)
        tax.append(t)
        net.append(ni)

    row_values = {
        ROW_REVENUE: revenue,
        ROW_COGS: cogs,
        ROW_GROSS_PROFIT: gross,
        ROW_OPEX: opex,
        ROW_OPERATING_INCOME: opinc,
        ROW_TAX: tax,
        ROW_NET_INCOME: net,
    }
    return years, row_values, sum(revenue)


def build_xlsx(task: ModelTask, a: Assumptions, out_path: str) -> str:
    """Build the workbook from assumptions using openpyxl. This is the EXECUTOR's
    deterministic half: every figure is derived here in code (via the loop's own
    `_derive_rows`), never supplied by the model. Crucially it does NOT call the
    verifier's `recompute` — the verifier re-derives the numbers from scratch with
    its own independent code and reconciles cell by cell, so this is a genuine
    cross-check, not the verifier trusting a copy of itself."""
    from openpyxl import Workbook  # local import: only needed when actually building

    years, row_values, total_revenue = _derive_rows(task, a)
    wb = Workbook()
    ws = wb.active
    ws.title = "Model"

    ws.cell(row=1, column=1, value=task.title)
    ws.cell(row=2, column=1, value=f"Currency: {task.currency}")

    # Year header.
    header_row = 4
    ws.cell(row=header_row, column=1, value="Year")
    for j, yr in enumerate(years):
        ws.cell(row=header_row, column=2 + j, value=yr)

    # Line items, one labelled row each, in canonical order.
    r = header_row + 1
    for label in ORDERED_ROWS:
        ws.cell(row=r, column=1, value=label)
        for j, v in enumerate(row_values[label]):
            ws.cell(row=r, column=2 + j, value=round(v, 2))
        r += 1

    # The cross-year Total Revenue cell (a balance identity the verifier checks).
    ws.cell(row=r + 1, column=1, value=ROW_TOTAL_REVENUE)
    ws.cell(row=r + 1, column=2, value=round(total_revenue, 2))

    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    wb.save(out_path)
    return out_path


class _ProposalSeed:
    """Nudges the generator toward a fresh assumption set each round so the loop
    explores rather than re-proposing the same numbers (which would dedupe out)."""

    def __init__(self) -> None:
        self.round = 0

    def hint(self) -> str:
        self.round += 1
        if self.round == 1:
            return "Propose a realistic SaaS-style projection."
        return (
            f"This is attempt #{self.round}. Propose a DIFFERENT, still-realistic "
            "assumption set than a typical default (vary growth and margins), so "
            "the loop explores new structures."
        )


def make_generator(llm: LLM, task: ModelTask, run_dir: str):
    """Return a `generate()` callable. Each call asks llm.strong for one fresh set
    of assumptions, builds the xlsx, and yields a single ModelCandidate."""
    seed = _ProposalSeed()

    def generate() -> Iterable[ModelCandidate]:
        assumptions = llm.structured(
            model=llm.strong,
            system=(
                "You are a financial analyst. Propose the ASSUMPTIONS for a "
                "multi-year operating model: base-year revenue, YoY revenue "
                "growth, COGS as a fraction of revenue, base-year operating "
                "expense, YoY opex growth, and a tax rate. Return only the "
                "assumptions — downstream code derives every projected figure. "
                "Keep figures realistic: margins sane, rates as decimals "
                "(0.15 == 15%)."
            ),
            user=(
                f"Task: {task.title}\n"
                f"Currency: {task.currency}\n"
                f"Years: {task.years} starting {task.start_year}\n"
                f"{seed.hint()}"
            ),
            schema=Assumptions,
        )
        candidate = ModelCandidate(task=task, assumptions=assumptions)
        out = os.path.join(run_dir, "models", f"{candidate.dedupe_key}.xlsx")
        candidate.xlsx_path = build_xlsx(task, assumptions, out)
        return [candidate]

    return generate


def build(
    run_dir: str,
    budget: Optional[Budget] = None,
    *,
    task: ModelTask = DEMO_TASK,
    llm: Optional[LLM] = None,
    max_rounds: int = 8,
):
    """Wire generate + verify + state + budget. Returns a zero-arg callable that
    runs the loop and returns a LoopResult."""
    budget = budget or Budget(max_usd=2.0)
    llm = llm or LLM(budget)
    state = JsonlState(run_dir)
    generate = make_generator(llm, task, run_dir)

    def run():
        return run_loop(
            generate=generate,
            verify=verify,  # deterministic — no llm passed in on purpose
            state=state,
            budget=budget,
            max_rounds=max_rounds,
            dry_rounds_to_stop=2,
        )

    return run


def main() -> None:
    run_dir = os.environ.get(
        "FINMODEL_RUN_DIR",
        os.path.join(os.path.dirname(__file__), "..", "..", ".runs", "finmodel"),
    )
    run_dir = os.path.abspath(run_dir)
    budget = Budget(max_usd=2.0)
    run = build(run_dir, budget)

    result = run()

    print("\n=== finmodel ===")
    print(f"rounds={result.rounds} stopped={result.stopped}")
    print(f"confirmed={len(result.confirmed)} rejected={len(result.rejected)}")
    print(f"budget: {budget.summary()}")
    for cand, verdict in result.confirmed:
        print(f"\nCONFIRMED {cand.dedupe_key}: {verdict.reason}")
        print(f"  xlsx: {verdict.evidence.get('xlsx_path')}")
        print(f"  recompute total_revenue: {verdict.evidence['recompute']['total_revenue']}")
    for cand, verdict in result.rejected:
        print(f"\nrejected {cand.dedupe_key}: {verdict.reason}")
        if verdict.evidence.get("reconcile_diffs"):
            print(f"  diffs: {verdict.evidence['reconcile_diffs'][:3]}")

    # HUMAN HANDOFF — plain-English summary of what you got and where the proof is.
    print("\n=== HUMAN HANDOFF ===")
    n = len(result.confirmed)
    if n:
        print(
            f"You got {n} financial model{'s' if n != 1 else ''} built as real "
            f".xlsx file{'s' if n != 1 else ''}. Every number in each one reconciled "
            "cell-by-cell against an independent Python recompute and passed every "
            "rubric criterion, so they're ready to send, not drafts to re-check by hand."
        )
        print("Open the spreadsheet(s):")
        for cand, verdict in result.confirmed:
            print(f"  - {verdict.evidence.get('xlsx_path')}")
    else:
        print(
            "No model reconciled clean this run — nothing is being handed off as "
            "confirmed. The evidence below shows exactly which cells disagreed so you "
            "can see why, rather than shipping a model that doesn't tie out."
        )
    print(
        "\nProof / evidence to verify the verdict yourself:\n"
        f"  - run directory: {run_dir}\n"
        f"  - confirmed evidence (per-candidate verdicts, reconcile diffs, rubric "
        f"booleans, recompute headline figures): {os.path.join(run_dir, 'confirmed.jsonl')}\n"
        f"  - rejected evidence (reasons + the cells that disagreed): "
        f"{os.path.join(run_dir, 'rejected.jsonl')}\n"
        "  - each verdict carries the recompute's headline figures and any "
        "reconciliation diffs, so a human can confirm the result in under a minute."
    )


if __name__ == "__main__":
    main()
